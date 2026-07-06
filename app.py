from flask import Flask, render_template, request, send_file
import sqlite3
from datetime import datetime, timedelta
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

app = Flask(__name__)

DB_PATH = "attendance.db"
FACES_DIR = "data/data_faces_from_camera"

GRACE_MINUTES = 5  # ✅ 5-minute grace period


# ===================== HELPERS =====================
def format_time_12h(time_str):
    return datetime.strptime(time_str, "%H:%M:%S").strftime("%I:%M %p")


def get_all_students():
    students = []
    if os.path.exists(FACES_DIR):
        for d in os.listdir(FACES_DIR):
            if d.startswith("person_"):
                students.append(d.split("_", 2)[-1])
    return sorted(students, key=str.lower)


def get_attendance_data(insertDate, class_start):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(
        "SELECT name, time FROM attendance WHERE date = ?",
        (date,)
    )
    records = cursor.fetchall()
    conn.close()

    attendance_map = {n: t for n, t in records}
    students = get_all_students()

    on_time, late, absent = [], [], []

    # class start + grace period
    class_start_dt = datetime.strptime(class_start, "%H:%M")
    grace_deadline = class_start_dt + timedelta(minutes=GRACE_MINUTES)

    for student in students:
        if student in attendance_map:
            time_in = attendance_map[student]
            time_in_dt = datetime.strptime(time_in, "%H:%M:%S")

            if time_in_dt <= grace_deadline:
                status = "On Time"
                on_time.append((student, format_time_12h(time_in), status))
            else:
                status = "Late"
                late.append((student, format_time_12h(time_in), status))
        else:
            absent.append((student, "-", "Absent"))

    return on_time, late, absent


# ===================== ROUTES =====================
@app.route("/")
def index():
    now = datetime.now()
    return render_template(
        "index.html",
        selected_date=now.strftime("%Y-%m-%d"),
        class_start=now.strftime("%H:%M")
    )


@app.route("/attendance", methods=["POST"])
def attendance():
    date = request.form["selected_date"]
    class_start = request.form["class_start"]

    on_time, late, absent = get_attendance_data(date, class_start)
    data = on_time + late + absent

    stats = {
        "on_time": len(on_time),
        "late": len(late),
        "absent": len(absent)
    }

    return render_template(
        "index.html",
        attendance_data=data,
        selected_date=date,
        class_start=class_start,
        total=len(data),
        stats=stats
    )


# ===================== EXCEL EXPORT =====================
@app.route("/export/excel", methods=["POST"])
def export_excel():
    date = request.form["date"]
    class_start = request.form["class_start"]

    on_time, late, absent = get_attendance_data(date, class_start)

    wb = Workbook()
    wb.remove(wb.active)

    def create_sheet(title, data):
        ws = wb.create_sheet(title)
        ws.append(["Name", "Time In", "Status"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in data:
            ws.append(row)

    create_sheet("On-Time", on_time)
    create_sheet("Late", late)
    create_sheet("Absent", absent)

    filename = "attendance.xlsx"
    wb.save(filename)

    return send_file(filename, as_attachment=True)


# ===================== PDF EXPORT =====================
@app.route("/export/pdf", methods=["POST"])
def export_pdf():
    date = request.form["date"]
    class_start = request.form["class_start"]

    on_time, late, absent = get_attendance_data(date, class_start)

    filename = "attendance.pdf"
    doc = SimpleDocTemplate(filename)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("<b>CAVITE STATE UNIVERSITY</b>", styles["Title"]))
    elements.append(Paragraph("Face Recognition Attendance Report", styles["Normal"]))
    elements.append(Paragraph(f"Date: {date}", styles["Normal"]))
    elements.append(Paragraph(f"Grace Period: {GRACE_MINUTES} minutes", styles["Normal"]))
    elements.append(Spacer(1, 15))

    def section(title, data, header_color):
        elements.append(Paragraph(f"<b>{title}</b>", styles["Heading2"]))
        table_data = [["Name", "Time In", "Status"]] + data
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), header_color),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))

    section("ON-TIME STUDENTS", on_time, colors.green)
    section("LATE STUDENTS", late, colors.orange)
    section("ABSENT STUDENTS", absent, colors.red)

    doc.build(elements)
    return send_file(filename, as_attachment=True)


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
